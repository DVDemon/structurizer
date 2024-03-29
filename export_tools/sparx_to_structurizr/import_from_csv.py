import openpyxl
import sys, getopt


def main(argv):
    # parse args
    inputfile = ''
    prefix = ''

    helpstring = 'import_from_csv.py -i <inputfile> -p <prefix>'
    try:
        opts, args = getopt.getopt(argv,"hp:i:o:",["ifile=","ofile="])
    except getopt.GetoptError:
        print (helpstring)
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print (helpstring)
            sys.exit()
        elif opt in ("-i"):
            inputfile = arg
        elif opt in ("-p"):
            prefix = arg


    if len(inputfile) == 0:
        print (helpstring)
        sys.exit()
    
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(inputfile)
    
    # Define variable to read sheet
    ws = dataframe.active
    
    # Iterate the loop to read the cell values
    ru_symbols = "абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ"

    print('# This is autogenerated file from Sparx Enterprise Application Catalogue')
    print('#')
    components = dict()
    for row in range(1, ws.max_row+1):

        if ws.cell(row=row, column=2).value == 'Component':
            cmdb_name = ws.cell(row=row, column=9).value

            if not (cmdb_name is None):
                has_russian = False
                for c in cmdb_name:
                    if c in ru_symbols:
                        has_russian = True
                        break
                
                if not has_russian:
                    if cmdb_name in components.keys():
                        components[cmdb_name].append(row)
                        print('# Found multiple elements for '+cmdb_name+' system')
                    else:
                        components[cmdb_name] = list()
                        components[cmdb_name].append(row)
                        
                else:
                    print('# System cmdb code has russian name: '+ws.cell(row=row, column=1).value+' code:'+cmdb_name)
            else:
                if not(ws.cell(row=row, column=9).value is None):
                    print('# No cmdb code for system name: '+ws.cell(row=row, column=1).value)


    symbols = [u"абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ abvgdeejzijklmnoprstufhzcssvw_y_euaABVGDEEJZIJKLMNOPRSTUFHZCSS_Y_EUA_VW",
            u"abvgdeejzijklmnoprstufhzcss_y_euaABVGDEEJZIJKLMNOPRSTUFHZCSS_Y_EUA_abvgdeejzijklmnoprstufhzcssvw_y_euaABVGDEEJZIJKLMNOPRSTUFHZCSS_Y_EUA_VW"]


    def print_system(system_name,title,cmdb_mnemonic,status,description,modified):
        print (system_name +' = softwareSystem "'+title+'" {')
        print ('   description "'+description+'"')
        print ('   tags "landscape" "'+prefix+'"')
        print ('   properties {')
        print ('    "cmdb" "'+cmdb_mnemonic+'"')
        print ('    "status" "'+status+'"')
        print ('    "domain" "'+prefix+'"')
        print ('    "modified" "'+modified+'"')
        print ('   }')
        print ('}')

    print()
    print()
    print()

    for cmdb_mnemonic in components.keys():
            system_name = ""
            system_name = cmdb_mnemonic.replace('.','_').replace(' ','_').replace('"','').replace('-','_')
                
            status = ""
            description   = ""
            modified = ""
            for row in components[cmdb_mnemonic]:
                if  not (ws.cell(row=row, column=17).value is None):
                    modified = ws.cell(row=row, column=17).value
                if  not (ws.cell(row=row, column=1).value is None):
                    description = ws.cell(row=row, column=1).value
                    description = description.replace('\n',' ').replace('"','').replace('\'','')
                if  not (ws.cell(row=row, column=10).value is None):
                    status = ws.cell(row=row, column=10).value
                    

            print_system(system_name,cmdb_mnemonic.replace('"','_'),cmdb_mnemonic,status,description,modified)
                

if __name__ == "__main__":
   main(sys.argv[1:])